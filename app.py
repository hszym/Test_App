import streamlit as st
import pandas as pd
import os
import time
import yfinance as yf
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, timedelta
from pathlib import Path

# --- NEW IMPORTS FOR AUTOMATION ---
import win32com.client as win32
import pythoncom

# --- CONFIGURATION ---
MASTER_EXCEL = r"C:\Users\Hugo Szym\OneDrive - Plurimi\Desktop\Mastersheet.xlsx"
BASE_TRADE_FOLDER = Path(r"C:\Users\Hugo Szym\OneDrive - Plurimi\Desktop\Trades")

DOC_TYPES = {"Client Confirmation": "Client_Confo", "Bank Confirmation": "Bank_Confo", "Termsheet": "TS"}
ALL_COLS = [
    "Deal ID", "Booker", "Client Name", "Issuer", "ISIN", "Sales",
    "Currency", "Nominal", "Exchange rate", "Retrocession", "Fee (%)",
    "Sales Credit (%)", "Traded Price", "Trade Time", "Trade Date",
    "Issue Date", "Maturity Date", "Redemption Date", "Product Type",
    "Underlying 1", "Underlying 2", "Underlying 3 & more",
    "Broker Used", "Broker Name", "Is Internal", "Expersoft ID"
]

COST_OF_SALES_ITEMS = [
    "Travel & Subsistence", "Entertainment (Client)", "Entertainment (Staff)",
    "Social Charges", "Pension Costs", "Medical Insurance", "IT Software",
    "Legal", "Research Fees", "Subscriptions", "Staff Benefits"
]

COSTS_COLS = ["Quarter", "Hugo_Salary", "Philip_Salary"] + [
    item.replace(" ", "_").replace("&", "and").replace("(", "").replace(")", "")
    for item in COST_OF_SALES_ITEMS
] + ["Open_Balance", "Paid_Bonus"]

COST_COL_MAP = {item: item.replace(" ", "_").replace("&", "and").replace("(", "").replace(")", "") for item in COST_OF_SALES_ITEMS}

CLIENTS_INFO_COLS  = ["Client Name", "Company", "Vizibility ID", "Expersoft ID", "Client Type", "Email", "Phone", "Mobile"]
SUB_CLIENTS_COLS   = ["Deal_ID", "Expersoft_ID", "Nominal"]
BANKING_INFO_COLS  = ["Company", "Currency", "Salutation", "Contact Name", "Address", "Postal Code", "Country", "Bank Name", "IBAN", "BIC", "Account Name"]
PAYMENTS_COLS      = ["Deal_ID", "Issuer", "Client Name", "ISIN", "Currency", "Nominal",
                      "Issuer Fee (%)", "Issuer Fee Amount",
                      "Retro Fee (%)", "Retro Amount",
                      "Issue Date", "Due Date",
                      "Issuer Status", "Issuer Received Date",
                      "Payment Status", "Payment Date"]

# --- DATA HELPERS ---

def fetch_rate_for_deal(currency, trade_date_str):
    date_obj = pd.to_datetime(trade_date_str)
    return get_exchange_rate(currency, date_obj)


def create_gauge(current, target):
    progress = (current / target * 100) if target > 0 else 0
    fig = go.Figure(go.Indicator(
        mode = "gauge+number",
        value = current,
        number = {'prefix': "€", 'font': {'size': 24}},
        title = {'text': "P&L Target Achievement", 'font': {'size': 18}},
        gauge = {
            'axis': {'range': [0, max(target, current * 1.2)], 'tickwidth': 1},
            'bar': {'color': "#1d3557"},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "gray",
            'steps': [
                {'range': [0, target], 'color': '#f1f4f9'},
                {'range': [target, max(target, current * 1.2)], 'color': '#e1f5fe'}],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': target}}))
    fig.update_layout(height=250, margin=dict(t=50, b=0, l=25, r=25))
    return fig

def load_data():
    if os.path.exists(MASTER_EXCEL):
        try:
            with pd.ExcelFile(MASTER_EXCEL) as xls:
                df = pd.read_excel(xls, sheet_name="Sheet1")
                costs_df        = pd.read_excel(xls, sheet_name="Costs")        if "Costs"        in xls.sheet_names else pd.DataFrame()
                contacts_df     = pd.read_excel(xls, sheet_name="Contacts")     if "Contacts"     in xls.sheet_names else pd.DataFrame()
                clients_info_df = pd.read_excel(xls, sheet_name="Clients_Info") if "Clients_Info" in xls.sheet_names else pd.DataFrame(columns=CLIENTS_INFO_COLS)
                sub_clients_df  = pd.read_excel(xls, sheet_name="Sub_Clients")  if "Sub_Clients"  in xls.sheet_names else pd.DataFrame(columns=SUB_CLIENTS_COLS)
                banking_info_df = pd.read_excel(xls, sheet_name="Banking_Info") if "Banking_Info" in xls.sheet_names else pd.DataFrame(columns=BANKING_INFO_COLS)
                # Migrate old column name (saved before rename)
                if 'Client Name' in banking_info_df.columns and 'Company' not in banking_info_df.columns:
                    banking_info_df = banking_info_df.rename(columns={'Client Name': 'Company'})
                payments_df     = pd.read_excel(xls, sheet_name="Payments")     if "Payments"     in xls.sheet_names else pd.DataFrame(columns=PAYMENTS_COLS)
                # Migrate old column names
                payments_df = payments_df.rename(columns={"Fee (%)": "Issuer Fee (%)", "Fee Amount": "Issuer Fee Amount"})
                for _mc in PAYMENTS_COLS:
                    if _mc not in payments_df.columns:
                        payments_df[_mc] = ''

            # Force text columns in payments_df to str so Streamlit data_editor doesn't see float NaN
            for _pc in ["Issuer Status", "Issuer Received Date", "Payment Status", "Payment Date"]:
                if _pc in payments_df.columns:
                    payments_df[_pc] = payments_df[_pc].fillna('').astype(str).replace('nan', '')

            # Safety: add missing columns to main blotter
            for c in [col for col in ALL_COLS if col not in df.columns]: df[c] = None

            return df.astype(object), costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df
        except Exception as e:
            st.error(f"Load error: {e}")
    return (pd.DataFrame(columns=ALL_COLS), pd.DataFrame(), pd.DataFrame(),
            pd.DataFrame(columns=CLIENTS_INFO_COLS), pd.DataFrame(columns=SUB_CLIENTS_COLS),
            pd.DataFrame(columns=BANKING_INFO_COLS), pd.DataFrame(columns=PAYMENTS_COLS))

def save_all_sheets(blotter_df, costs_df, contacts_df, clients_info_df=None,
                    sub_clients_df=None, banking_info_df=None, payments_df=None):
    if clients_info_df  is None: clients_info_df  = pd.DataFrame(columns=CLIENTS_INFO_COLS)
    if sub_clients_df   is None: sub_clients_df   = pd.DataFrame(columns=SUB_CLIENTS_COLS)
    if banking_info_df  is None: banking_info_df  = pd.DataFrame(columns=BANKING_INFO_COLS)
    if payments_df      is None: payments_df      = pd.DataFrame(columns=PAYMENTS_COLS)
    try:
        settings_df = pd.read_excel(MASTER_EXCEL, sheet_name="Data Settings")
        with pd.ExcelWriter(MASTER_EXCEL, engine='openpyxl') as writer:
            blotter_df.to_excel(writer,      sheet_name="Sheet1",       index=False)
            settings_df.to_excel(writer,     sheet_name="Data Settings", index=False)
            costs_df.to_excel(writer,        sheet_name="Costs",         index=False)
            contacts_df.to_excel(writer,     sheet_name="Contacts",      index=False)
            clients_info_df.to_excel(writer, sheet_name="Clients_Info",  index=False)
            sub_clients_df.to_excel(writer,  sheet_name="Sub_Clients",   index=False)
            banking_info_df.to_excel(writer, sheet_name="Banking_Info",  index=False)
            payments_df.to_excel(writer,     sheet_name="Payments",      index=False)
        return True
    except Exception as e:
        st.error(f"Save Error: {e}")
        return False

def get_exchange_rate(ccy, trade_date_obj):
    """Return CCY/EUR rate (how many EUR per 1 CCY unit) for the given trade date.
    Primary: ECB SDW API — EXR D.{CCY}.EUR.SP00.A gives CCY per EUR, so we invert.
    Fallback: yfinance.
    """
    if ccy == "EUR": return 1.0
    import urllib.request, csv, io
    dt = pd.to_datetime(trade_date_obj)
    # Look up to 14 calendar days back so we always find a business day
    start = (dt - timedelta(days=14)).strftime('%Y-%m-%d')
    end   = dt.strftime('%Y-%m-%d')

    # --- ECB API ---
    try:
        url = (
            f"https://data-api.ecb.europa.eu/service/data/EXR/"
            f"D.{ccy}.EUR.SP00.A"
            f"?startPeriod={start}&endPeriod={end}&format=csvdata"
        )
        req = urllib.request.Request(url, headers={"Accept": "text/csv"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            content = resp.read().decode("utf-8")
        reader = csv.DictReader(io.StringIO(content))
        obs_values = [
            float(row["OBS_VALUE"])
            for row in reader
            if row.get("OBS_VALUE", "").strip()
        ]
        if obs_values:
            # ECB EXR D.{CCY}.EUR.SP00.A gives CCY per EUR (e.g. USDEUR=1.16 means 1 EUR = 1.16 USD)
            # We need EUR per CCY, so invert.
            return round(1 / obs_values[-1], 4)
    except Exception as ecb_err:
        ecb_msg = str(ecb_err)
    else:
        ecb_msg = "no data"

    # --- yfinance fallback ---
    yf_start = (dt - timedelta(days=10)).strftime('%Y-%m-%d')
    yf_end   = (dt + timedelta(days=1)).strftime('%Y-%m-%d')
    def _fetch(ticker):
        try:
            data = yf.Ticker(ticker).history(start=yf_start, end=yf_end)
            if not data.empty:
                return round(float(data['Close'].iloc[-1]), 4)
        except Exception:
            pass
        return None

    val = _fetch(f"{ccy}EUR=X")
    if val: return val
    val = _fetch(f"EUR{ccy}=X")
    if val: return round(1 / val, 4)

    raise ValueError(
        f"No FX data found for {ccy}/EUR. "
        f"ECB: {ecb_msg}. "
        f"yfinance also failed for {ccy}EUR=X and EUR{ccy}=X ({yf_start}→{yf_end})."
    )

def find_deal_folder(deal_id):
    if not BASE_TRADE_FOLDER.exists(): return None
    for entry in BASE_TRADE_FOLDER.iterdir():
        if entry.is_dir() and entry.name.startswith(str(deal_id)):
            return entry
    return None

def get_deal_path(deal_id, row):
    """Return existing deal folder, or build the correct-convention path if none exists."""
    existing = find_deal_folder(deal_id)
    if existing:
        return existing
    try:
        td_str = pd.to_datetime(row.get('Trade Date', '')).strftime('%d.%m.%Y')
    except:
        td_str = str(row.get('Trade Date', ''))
    nom_val = float(row.get('Nominal') or 0)
    nom_str = f"{int(nom_val):,}" if nom_val else "0"
    curr = row.get('Currency', '')
    isin = row.get('ISIN', '')
    return BASE_TRADE_FOLDER / f"{deal_id}_{td_str}_{isin}_{nom_str} {curr}"

def check_docs(deal_id):
    result = {slug: False for slug in DOC_TYPES.values()}
    deal_path = find_deal_folder(deal_id)
    if deal_path and deal_path.exists():
        files = os.listdir(deal_path)
        for slug in DOC_TYPES.values():
            if any(slug in f for f in files):
                result[slug] = True
    return result

def save_settings_to_excel(settings_dict):
    try:
        max_len = max(len(v) for v in settings_dict.values()) if settings_dict else 0
        padded = {k: v + [None] * (max_len - len(v)) for k, v in settings_dict.items()}
        new_settings_df = pd.DataFrame(padded)
        blotter_df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df = load_data()
        # Write settings sheet manually, then delegate the rest to save_all_sheets
        with pd.ExcelWriter(MASTER_EXCEL, engine='openpyxl') as writer:
            blotter_df.to_excel(writer,      sheet_name="Sheet1",        index=False)
            new_settings_df.to_excel(writer, sheet_name="Data Settings", index=False)
            costs_df.to_excel(writer,        sheet_name="Costs",         index=False)
            contacts_df.to_excel(writer,     sheet_name="Contacts",      index=False)
            clients_info_df.to_excel(writer, sheet_name="Clients_Info",  index=False)
            sub_clients_df.to_excel(writer,  sheet_name="Sub_Clients",   index=False)
            banking_info_df.to_excel(writer, sheet_name="Banking_Info",  index=False)
            payments_df.to_excel(writer,     sheet_name="Payments",      index=False)
        return True
    except Exception as e:
        st.error(f"Settings save error: {e}")
        return False

def reset_form():
    st.session_state.form_reset_key = st.session_state.get("form_reset_key", 0) + 1

# --- P&L EXCEL EXPORT ---
def build_pl_export(df, costs_df):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    REV_LINES = [
        'Arrangement Fees', 'AMC Revenues', 'Management Fees (Advisory)',
        'Transactional Retrocession Fees', 'Structured Product Revenue',
        'Inter Company Revenues', 'Sub-distribution fees',
    ]

    def rev_cat(row):
        is_int = str(row.get('Is Internal', '')).strip().lower() == 'yes'
        retro  = str(row.get('Retrocession', '')).strip().lower() == 'yes'
        pt     = str(row.get('Product Type', '')).strip().lower()
        if is_int:                                return 'Inter Company Revenues'
        if retro:                                 return 'Transactional Retrocession Fees'
        if 'amc' in pt:                           return 'AMC Revenues'
        if 'management' in pt or 'advisory' in pt: return 'Management Fees (Advisory)'
        if 'arrangement' in pt:                   return 'Arrangement Fees'
        if 'fund' in pt:                          return 'Fund Fees'
        if 'sub' in pt and 'distrib' in pt:       return 'Sub-distribution fees'
        return 'Structured Product Revenue'

    def who(row):
        for field in ['Sales', 'Booker']:
            n = str(row.get(field, '')).lower()
            if 'hugo'   in n: return 'Hugo'
            if 'philip' in n: return 'Philip'
        return 'Other'

    tdf = df.copy()
    tdf['_d']   = pd.to_datetime(tdf['Trade Date'], format='%d-%b-%y', errors='coerce')
    tdf         = tdf.dropna(subset=['_d'])
    tdf['_q']   = tdf['_d'].dt.to_period('Q').astype(str)
    tdf['_nom'] = pd.to_numeric(tdf['Nominal'],          errors='coerce').fillna(0)
    tdf['_sc']  = pd.to_numeric(tdf['Sales Credit (%)'], errors='coerce').fillna(0)
    tdf['_fx']  = pd.to_numeric(tdf['Exchange rate'],    errors='coerce').fillna(1)
    tdf['_rev'] = tdf['_nom'] * (tdf['_sc'] / 100) * tdf['_fx']
    tdf['_cat'] = tdf.apply(rev_cat, axis=1)
    tdf['_who'] = tdf.apply(who, axis=1)

    quarters = sorted(tdf['_q'].unique())
    if not quarters:
        buf = io.BytesIO(); Workbook().save(buf); buf.seek(0); return buf.getvalue()

    wb = Workbook()

    # ── Styles ────────────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="1D3557")
    hdr_font = Font(bold=True, color="FFFFFF")
    sub_fill = PatternFill("solid", fgColor="E8ECF2")
    tot_fill = PatternFill("solid", fgColor="457B9D")
    tot_font = Font(bold=True, color="FFFFFF")
    bold_f   = Font(bold=True)
    ital_f   = Font(italic=True)
    num_fmt  = '#,##0'
    center   = Alignment(horizontal='center')
    right_al = Alignment(horizontal='right')

    # ── Sheet 1: P&L Summary ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "P&L Summary"
    ws['A1'] = 'Plurimi Wealth Monaco S.A.M.'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = f'Investment Solutions — Quarterly Revenue & P&L  |  Generated {datetime.today().strftime("%d/%m/%Y")}'
    ws['A2'].font = Font(size=10, color="888888")
    ws.column_dimensions['A'].width = 38

    HDR, SUB, R0 = 4, 5, 6
    ws.cell(HDR, 1, 'Investment Solutions').font = bold_f

    col_start = {}
    c = 2
    for q in quarters:
        ws.cell(HDR, c, q).fill      = hdr_fill
        ws.cell(HDR, c).font         = hdr_font
        ws.cell(HDR, c).alignment    = center
        ws.merge_cells(start_row=HDR, start_column=c, end_row=HDR, end_column=c+2)
        for i, lbl in enumerate(['Hugo Szym', 'Philip Wingeier', 'Total']):
            cell = ws.cell(SUB, c+i, lbl)
            cell.fill = sub_fill; cell.font = bold_f; cell.alignment = center
            ws.column_dimensions[get_column_letter(c+i)].width = 16
        col_start[q] = c
        c += 3

    def sec_hdr(r, label):
        ws.cell(r, 1, label).font = hdr_font
        ws.cell(r, 1).fill = hdr_fill
        for q in quarters:
            for off in range(3):
                ws.cell(r, col_start[q]+off).fill = hdr_fill
        return r + 1

    def write_line(r, label, vals, style='normal'):
        cell = ws.cell(r, 1, label)
        if   style == 'total': cell.font = tot_font; cell.fill = tot_fill
        elif style == 'sub':   cell.font = bold_f;   cell.fill = sub_fill
        else:                  cell.font = ital_f
        for q in quarters:
            h = vals.get(q, {}).get('Hugo',   0) or 0
            p = vals.get(q, {}).get('Philip', 0) or 0
            t = h + p
            for off, v in enumerate([h, p, t]):
                obj = ws.cell(r, col_start[q]+off, round(v))
                obj.number_format = num_fmt; obj.alignment = right_al
                if   style == 'total': obj.font = tot_font; obj.fill = tot_fill
                elif style == 'sub':   obj.fill = sub_fill
        return r + 1

    row = R0

    # Gross Revenues
    row = sec_hdr(row, 'Gross Revenues')
    rev_totals = {q: {'Hugo': 0, 'Philip': 0} for q in quarters}
    for cat in REV_LINES:
        cat_vals = {}
        for q in quarters:
            mask = (tdf['_q'] == q) & (tdf['_cat'] == cat)
            h = tdf[mask & (tdf['_who'] == 'Hugo')  ]['_rev'].sum()
            p = tdf[mask & (tdf['_who'] == 'Philip')]['_rev'].sum()
            o = tdf[mask & (tdf['_who'] == 'Other') ]['_rev'].sum()
            cat_vals[q] = {'Hugo': h + o/2, 'Philip': p + o/2}
            rev_totals[q]['Hugo']   += h + o/2
            rev_totals[q]['Philip'] += p + o/2
        row = write_line(row, f'    {cat}', cat_vals)
    row = write_line(row, 'Total Gross Revenues', rev_totals, 'total')
    row += 1

    # Cost of Sales
    row = sec_hdr(row, 'Cost of Sales')
    cos_totals = {q: {'Hugo': 0, 'Philip': 0} for q in quarters}
    for item in COST_OF_SALES_ITEMS:
        slug = COST_COL_MAP[item]
        item_vals = {}
        for q in quarters:
            h = p = 0.0
            if not costs_df.empty:
                qr = costs_df[costs_df['Quarter'] == q]
                if not qr.empty:
                    h_col = f'{slug}_Hugo';  p_col = f'{slug}_Philip'
                    h = float(qr[h_col].fillna(0).iloc[0]) if h_col in qr.columns else 0.0
                    p = float(qr[p_col].fillna(0).iloc[0]) if p_col in qr.columns else 0.0
            item_vals[q] = {'Hugo': h, 'Philip': p}
            cos_totals[q]['Hugo']   += h
            cos_totals[q]['Philip'] += p
        row = write_line(row, f'    {item}', item_vals)
    row = write_line(row, 'Total Cost of Sales', cos_totals, 'total')
    row += 1

    # Net Revenue
    net = {q: {'Hugo':   rev_totals[q]['Hugo']   - cos_totals[q]['Hugo'],
               'Philip': rev_totals[q]['Philip'] - cos_totals[q]['Philip']} for q in quarters}
    row = write_line(row, 'Net Revenue', net, 'total')
    row += 1

    # Salary & Bonus
    row = sec_hdr(row, 'Salary & Bonus')
    sal = {}
    for q in quarters:
        h = p = 0.0
        if not costs_df.empty:
            qr = costs_df[costs_df['Quarter'] == q]
            if not qr.empty:
                h = float(qr['Hugo_Salary'].fillna(0).iloc[0])   if 'Hugo_Salary'   in qr.columns else 0.0
                p = float(qr['Philip_Salary'].fillna(0).iloc[0]) if 'Philip_Salary' in qr.columns else 0.0
        sal[q] = {'Hugo': -h, 'Philip': -p}
    row = write_line(row, '    Salary', sal)
    row = write_line(row, 'Total Salary & Bonus', sal, 'total')
    row += 1

    # Total Compensation (40%) = Net Revenue × 40%
    comp = {q: {'Hugo': net[q]['Hugo'] * 0.4, 'Philip': net[q]['Philip'] * 0.4} for q in quarters}
    row = write_line(row, 'Total Compensation (40%)', comp, 'total')

    # Net Compensation = Compensation - Salary
    net_comp = {q: {'Hugo':   comp[q]['Hugo']   + sal[q]['Hugo'],
                    'Philip': comp[q]['Philip'] + sal[q]['Philip']} for q in quarters}
    row = write_line(row, 'Net Compensation', net_comp, 'total')

    # ── Sheet 2: Raw Trades ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Trades")
    export_cols = [col for col in df.columns if not str(col).startswith('_')]
    for ci, col_name in enumerate(export_cols, 1):
        ws2.cell(1, ci, col_name).font = bold_f
        ws2.column_dimensions[get_column_letter(ci)].width = 18
    for ri, (_, row_data) in enumerate(df[export_cols].iterrows(), 2):
        for ci, val in enumerate(row_data, 1):
            ws2.cell(ri, ci, '' if pd.isna(val) else val)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# --- INVOICE PDF GENERATION ---
from datetime import timedelta
import pandas as pd

def _add_business_days(dt, days):
    """Add `days` business days to a datetime, skipping weekends."""
    count = 0
    while count < days:
        dt += timedelta(days=1)
        if dt.weekday() < 5:
            count += 1
    return dt

def draft_invoice_pdf(deal_row, banking_info_df, deal_path, clients_info_df=None):
    """Generate a PDF invoice for an external-retro deal and save to deal_path.
    Uses Word COM automation (already available via win32com).
    Returns the Path to the generated PDF, or raises on failure.
    """
    row = deal_row if isinstance(deal_row, dict) else deal_row.to_dict()
    client_name = str(row.get('Client Name', ''))
    curr        = str(row.get('Currency', ''))

    # Resolve company from clients_info_df
    company = client_name  # fallback
    if clients_info_df is not None and not clients_info_df.empty:
        ci = clients_info_df[clients_info_df['Client Name'] == client_name]
        if not ci.empty:
            company = str(ci.iloc[0].get('Company', '') or client_name)
    nom         = float(row.get('Nominal') or 0)
    fee_pct     = float(row.get('Fee (%)') or 0)
    pay_amount  = round(nom * fee_pct / 100, 2)
    issuer      = str(row.get('Issuer', ''))
    isin        = str(row.get('ISIN', ''))
    sales       = str(row.get('Sales', ''))

    try:
        issue_dt  = pd.to_datetime(row.get('Issue Date'))
        value_dt  = _add_business_days(issue_dt, 2)
        value_date_str = value_dt.strftime('%d %B %Y')
        trade_date_str = pd.to_datetime(row.get('Trade Date')).strftime('%d %B %Y')
    except Exception:
        value_date_str = str(row.get('Issue Date', ''))
        trade_date_str = str(row.get('Trade Date', ''))

    today_str = datetime.today().strftime('%d %B %Y')

    # --- Lookup banking info by Company + Currency ---
    bk = pd.DataFrame()
    if banking_info_df is not None and not banking_info_df.empty:
        bk = banking_info_df[
            (banking_info_df['Company'].astype(str) == company) &
            (banking_info_df['Currency'].astype(str).str.upper() == curr.upper())
        ]
        if bk.empty:
            # Fallback: any row for this company regardless of currency
            bk = banking_info_df[banking_info_df['Company'].astype(str) == company]
    if not bk.empty:
        b = bk.iloc[0]
        salutation   = str(b.get('Salutation',   '') or '')
        contact_name = str(b.get('Contact Name', '') or '')
        address      = str(b.get('Address',      '') or '')
        postal_code  = str(b.get('Postal Code',  '') or '')
        country      = str(b.get('Country',      '') or '')
        bank_name    = str(b.get('Bank Name',    '') or '')
        iban         = str(b.get('IBAN',         '') or '')
    else:
        salutation = contact_name = address = postal_code = country = bank_name = iban = ''

    nom_fmt = f"{nom:,.0f}"
    pay_fmt = f"{pay_amount:,.2f}"

    import base64, tempfile, subprocess
    from string import Template

    # Embed logo as base64
    logo_b64 = ""
    logo_path = Path(r"C:\Users\Hugo Szym\OneDrive - Plurimi\Desktop\Plurimi logo.png")
    if logo_path.exists():
        logo_b64 = base64.b64encode(logo_path.read_bytes()).decode()
    logo_tag = (
        f'<img src="data:image/png;base64,{logo_b64}" style="height:36px;" alt="Plurimi">'
        if logo_b64 else
        '<span style="font-size:20pt;font-weight:bold;color:#fff;letter-spacing:2px;">PLURIMI</span>'
    )

    # Load and fill template
    template_path = Path(r"C:\Users\Hugo Szym\OneDrive - Plurimi\Desktop\invoice_template.html")
    html = Template(template_path.read_text(encoding="utf-8")).substitute(
        NAVY="#1a2b4a",
        LIGHT_NAVY="#e8ecf2",
        logo_tag=logo_tag,
        isin=isin, issuer=issuer, curr=curr, sales=sales,
        company=company, salutation=salutation, contact_name=contact_name,
        address=address, postal_code=postal_code, country=country,
        today_str=today_str, trade_date_str=trade_date_str, value_date_str=value_date_str,
        nom_fmt=nom_fmt, pay_fmt=pay_fmt, fee_pct=fee_pct,
        bank_name=bank_name, iban=iban,
    )

    # Output PDF path
    deal_path = Path(deal_path)
    deal_path.mkdir(parents=True, exist_ok=True)
    pdf_path = deal_path / f"Invoice_{row.get('Deal ID','')}_{isin}.pdf"

    # Write temp HTML
    tmp_html = Path(tempfile.gettempdir()) / f"invoice_{row.get('Deal ID', 'tmp')}_{int(datetime.now().timestamp())}.html"
    tmp_html.write_text(html, encoding='utf-8')

    # Find Edge or Chrome (both support headless PDF printing)
    browser = None
    for candidate in [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    ]:
        if Path(candidate).exists():
            browser = candidate
            break

    if not browser:
        tmp_html.unlink(missing_ok=True)
        raise RuntimeError("No headless browser found (Edge or Chrome). Please install one.")

    subprocess.run([
        browser,
        "--headless=new",
        "--disable-gpu",
        "--no-sandbox",
        "--no-pdf-header-footer",
        f"--print-to-pdf={pdf_path.resolve()}",
        tmp_html.as_uri(),
    ], check=True, capture_output=True)

    tmp_html.unlink(missing_ok=True)
    return pdf_path


# --- AUTOMATION LOGIC (MIFIR + OUTLOOK) ---

def finalize_and_send(row_idx, df, settings_df, clients_info_df=None, sub_clients_df=None,
                      banking_info_df=None, payments_df=None):
    pythoncom.CoInitialize()
    row = df.loc[row_idx].to_dict()
    did = str(row.get('Deal ID', ''))
    is_internal = str(row.get('Is Internal', '')).strip().lower() == "yes"
    is_retrocession = str(row.get('Retrocession', '')).strip().lower() == "yes"
    issuer_name = str(row.get('Issuer', ''))
    client_name = str(row.get('Client Name', ''))

    # 1. Client type lookup
    ci_match = clients_info_df[clients_info_df['Client Name'] == client_name] if clients_info_df is not None and not clients_info_df.empty else pd.DataFrame()
    client_type = str(ci_match.iloc[0].get('Client Type', 'External')).strip() if not ci_match.empty else 'External'
    client_expersoft_id = str(ci_match.iloc[0].get('Expersoft ID', '')) if not ci_match.empty else ''

    # 2. Economics
    nom = float(row.get('Nominal') or 0)
    fee_pct = float(row.get('Fee (%)') or 0)
    sc_pct = float(row.get('Sales Credit (%)') or 0)
    curr = row.get('Currency', '')
    fee_val = round(nom * (fee_pct / 100), 2)
    sc_val = round(nom * (sc_pct / 100), 2)
    total_claim = fee_val + sc_val

    # 3. Receipt Date
    issue_date_raw = row.get('Issue Date')
    try:
        issue_dt = pd.to_datetime(issue_date_raw)
        due_date = issue_dt + timedelta(days=2)
        due_date_str = due_date.strftime('%d %b %Y')
    except:
        due_date_str = "T+2 from Issue Date"

    # 4. Issuer settings lookup
    match = settings_df[settings_df['Issuer'] == issuer_name]
    s = match.iloc[0].to_dict() if not match.empty else {}
    receiving_entity = s.get('Receiving Entity', 'NOT SET')
    cust_id = s.get('Customer ID', '')
    c1 = s.get('Contact1', '')
    c2 = s.get('Contact2', '')
    c3 = s.get('Contact3', '')

    # 5. MIFIR Excel Generation
    mifir_columns = [
        "Portfolio number (as per Expersoft)", "SP Fee", "Transaction Id (see Tab 2)",
        "Transaction type (See Tab 2)", "Transaction date & Time", "Transaction time",
        "Value date", "ISIN", "Currency", "Amount", "Price", "Account currency",
        "Fee Type", "Plurimi Fee (Value)", "Fee Type", "Issuer Fee (Value)",
        "Contact Details", "Issuer Name", "Customer ID",
        "Invoice Contact 1", "Invoice Contact 2", "Invoice Contact 3"
    ]

    mifir_data = []
    def create_mifir_row(port_id, val, amount=None):
        return [
            port_id, "Fee paid by Issuer", 1, "BUY",
            row.get('Trade Date', ''), row.get('Trade Time', ''), issue_date_raw,
            row.get('ISIN', ''), curr, amount if amount is not None else nom, row.get('Traded Price', ''),
            curr, "Plurimi One Off", val, "Financial Instrument One Off", 0,
            "Please provide contact details", issuer_name, cust_id,
            c1, c2, c3
        ]

    if client_type == 'Plurimi':
        # One MIFIR row per sub-client — filter out empty rows
        deal_subs = sub_clients_df[sub_clients_df['Deal_ID'] == did] if sub_clients_df is not None and not sub_clients_df.empty else pd.DataFrame()
        if not deal_subs.empty:
            deal_subs = deal_subs[
                deal_subs['Expersoft_ID'].notna() &
                (deal_subs['Expersoft_ID'].astype(str).str.strip() != "") &
                (pd.to_numeric(deal_subs['Nominal'], errors='coerce').fillna(0) > 0)
            ]
        for _, sub in deal_subs.iterrows():
            sub_nom = float(sub.get('Nominal', 0))
            sub_fee = round(sub_nom * (fee_pct / 100), 2)
            mifir_data.append(create_mifir_row(str(sub.get('Expersoft_ID', '')), sub_fee, sub_nom))
        # One extra row: no Expersoft ID, full nominal, sales credit only
        mifir_data.append(create_mifir_row('', sc_val, nom))
    elif client_type == 'Individual':
        mifir_data.append(create_mifir_row(client_expersoft_id, total_claim))
    else:
        # External — original logic
        if is_internal:
            mifir_data.append(create_mifir_row(row.get('Expersoft ID', ''), fee_val))
            mifir_data.append(create_mifir_row("", sc_val))
        else:
            mifir_data.append(create_mifir_row(row.get('Expersoft ID', ''), total_claim))

    # Save MIFIR Excel
    deal_path = get_deal_path(did, row)
    os.makedirs(deal_path, exist_ok=True)
    mifir_file = deal_path / f"MIFIR_{did}_{row.get('ISIN', '')}.xlsx"
    pd.DataFrame(mifir_data, columns=mifir_columns).to_excel(mifir_file, index=False)

    # 6. Email — build client-type-specific rows
    td_style = 'style="background-color: #f2f2f2; width: 220px;"'

    if client_type == 'Plurimi':
        deal_subs = sub_clients_df[sub_clients_df['Deal_ID'] == did] if sub_clients_df is not None and not sub_clients_df.empty else pd.DataFrame()
        if not deal_subs.empty:
            deal_subs = deal_subs[
                deal_subs['Expersoft_ID'].notna() &
                (deal_subs['Expersoft_ID'].astype(str).str.strip() != "") &
                (pd.to_numeric(deal_subs['Nominal'], errors='coerce').fillna(0) > 0)
            ]
        sub_lines = "".join(
            f"ID: <b>{row2['Expersoft_ID']}</b> — {float(row2['Nominal']):,.0f} {curr}<br>"
            for _, row2 in deal_subs.iterrows()
        ) if not deal_subs.empty else "No sub-clients recorded."
        extra_rows = f"""
                <tr><td {td_style}><b>Sub-Client Breakdown</b></td><td>{sub_lines}</td></tr>
                <tr><td {td_style}><b>Allocation</b></td><td>
                    <b>{client_name}</b>: {fee_val:,.2f} {curr} ({fee_pct}%)<br>
                    <b>{row.get('Sales', 'Sales')}</b>: {sc_val:,.2f} {curr} ({sc_pct}%)
                </td></tr>"""
    elif client_type == 'Individual':
        extra_rows = f"""
                <tr><td {td_style}><b>Client Expersoft ID</b></td><td>{client_expersoft_id}</td></tr>"""
    else:
        extra_rows = f"""
                <tr>
                    <td {td_style}><b>Breakdown</b></td>
                    <td>
                        {"<b>" + client_name + "</b>: " + f"{fee_val:,.2f} {curr} ({fee_pct}%)<br>" if is_retrocession else ""}
                        <b>{row.get('Sales', 'Sales')}</b>: {sc_val:,.2f} {curr} ({sc_pct}%)
                    </td>
                </tr>"""

    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    try:
        _td_fmt = pd.to_datetime(row.get('Trade Date', '')).strftime('%d-%b-%Y')
    except Exception:
        _td_fmt = str(row.get('Trade Date', ''))
    mail.Subject = f"Trade Confirmation: {did}_{_td_fmt}_{row.get('ISIN', '')}_{nom:,.0f} {curr}"
    mail.HTMLBody = f"""
    <html>
        <body style="font-family: Calibri, sans-serif;">
            <p>Hi Teams,</p>
            <p>Please find below details on the trade ref <b>{did}</b>.</p>
            <table border="1" cellpadding="5" style="border-collapse: collapse; width: 650px;">
                <tr><td {td_style}><b>Client</b></td><td>{client_name}</td></tr>
                <tr><td {td_style}><b>Trading Entity</b></td><td>{issuer_name}</td></tr>
                <tr><td {td_style}><b>Amount to be claimed</b></td><td><b>{total_claim:,.2f} {curr}</b> ({fee_pct + sc_pct}% of {nom:,.0f})</td></tr>
                <tr><td {td_style}><b>Expected Receipt Date</b></td><td><b>{due_date_str}</b></td></tr>
                <tr><td {td_style}><b>Receiving Entity</b></td><td>{receiving_entity}</td></tr>
                {extra_rows}
            </table>
            <br>
            {"<p>" + client_name + " and his team will share the supporting documentation for each sub-client's confirmation.</p>" if client_type == 'Plurimi' else ""}
            <p><i>Invoice and Bill have NOT been submitted for approval in Xero.</i></p>
            <p>Best regards,<br>Middle Office Hub</p>
        </body>
    </html>
    """
    mail.Attachments.Add(str(mifir_file))

    # Attach Client Confirmation and Bank (Issuer) Confirmation if present
    for slug in ["Client_Confo", "Bank_Confo"]:
        try:
            all_files = os.listdir(deal_path)
            match_file = next((f for f in all_files if slug in f), None)
            if match_file:
                mail.Attachments.Add(str(deal_path / match_file))
        except Exception:
            pass

    mail.Display()

    # 7. Auto-create / update payment record for ALL deals
    if payments_df is None:
        payments_df = pd.DataFrame(columns=PAYMENTS_COLS)
    try:
        issue_dt_pay   = pd.to_datetime(issue_date_raw)
        due_dt_pay     = _add_business_days(issue_dt_pay, 2)
        due_date_pay   = due_dt_pay.strftime('%d-%b-%y')
        issue_date_pay = issue_dt_pay.strftime('%d-%b-%y')
    except Exception:
        due_date_pay   = str(issue_date_raw)
        issue_date_pay = str(issue_date_raw)

    new_record = {
        "Deal_ID":               did,
        "Issuer":                issuer_name,
        "Client Name":           client_name,
        "ISIN":                  row.get('ISIN', ''),
        "Currency":              curr,
        "Nominal":               nom,
        "Issuer Fee (%)":        fee_pct + sc_pct,
        "Issuer Fee Amount":     total_claim,
        "Retro Fee (%)":         fee_pct if (client_type == 'External' and is_retrocession) else 0,
        "Retro Amount":          fee_val  if (client_type == 'External' and is_retrocession) else 0,
        "Issue Date":            issue_date_pay,
        "Due Date":              due_date_pay,
        "Issuer Status":         "Pending",
        "Issuer Received Date":  "",
        "Payment Status":        "Pending" if (client_type == 'External' and is_retrocession) else "N/A",
        "Payment Date":          "",
    }
    # Replace existing record for this deal, or append
    if did in payments_df['Deal_ID'].astype(str).values:
        payments_df = payments_df[payments_df['Deal_ID'].astype(str) != did]
    payments_df = pd.concat([payments_df, pd.DataFrame([new_record])], ignore_index=True)

    return payments_df

# --- SESSION STATE INIT ---
if "active_deal" not in st.session_state: st.session_state.active_deal = None
if "form_reset_key" not in st.session_state: st.session_state.form_reset_key = 0

# --- UI SETUP ---
st.set_page_config(page_title="Middle Office Hub", layout="wide")
df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df = load_data()
try:
    settings_df = pd.read_excel(MASTER_EXCEL, sheet_name="Data Settings")
    settings = {col: settings_df[col].dropna().astype(str).tolist() for col in settings_df.columns}
except:
    settings = {}

tab_perf, tab_dash, tab_book, tab_costs, tab_pay, tab_settings = st.tabs([
    "📈 Performance", "🎯 Action Dashboard", "➕ New Booking", "💸 Costs Management", "💳 Payments", "⚙️ Settings"
])

# --- TAB: PERFORMANCE ---
with tab_perf:
    if not df.empty:
        pdf = df.copy()
        pdf['Trade Date'] = pd.to_datetime(pdf['Trade Date'], format='%d-%b-%y', errors='coerce')
        pdf = pdf.dropna(subset=['Trade Date'])
        # Compute revenue per trade: SC% × Nominal × FX → EUR
        pdf['Nominal_n'] = pd.to_numeric(pdf['Nominal'], errors='coerce').fillna(0)
        pdf['SC_n']      = pd.to_numeric(pdf['Sales Credit (%)'], errors='coerce').fillna(0)
        pdf['FX_n']      = pd.to_numeric(pdf['Exchange rate'], errors='coerce').fillna(1)
        pdf['Rev_EUR']   = pdf['Nominal_n'] * (pdf['SC_n'] / 100) * pdf['FX_n']
        pdf['Quarter']   = pdf['Trade Date'].dt.to_period('Q').astype(str)

        q_list = sorted(pdf['Quarter'].unique(), reverse=True)
        sel_q = st.selectbox("📅 Filter View", ["All Time"] + q_list, index=1 if q_list else 0)
        f_pdf = pdf if sel_q == "All Time" else pdf[pdf['Quarter'] == sel_q]

        total_rev = f_pdf['Rev_EUR'].sum()

        # YTD: fiscal year runs April → March
        now = datetime.now()
        fy_start = datetime(now.year if now.month >= 4 else now.year - 1, 4, 1)
        ytd_pdf = pdf[pdf['Trade Date'] >= pd.Timestamp(fy_start)]
        ytd_rev = ytd_pdf['Rev_EUR'].sum()

        # Quarter costs
        total_salaries = 0; total_cos = 0; open_balance = 0; paid_bonus = 0; target = 0
        if not costs_df.empty and sel_q != "All Time":
            c_row = costs_df[costs_df['Quarter'] == sel_q]
            if not c_row.empty:
                row = c_row.iloc[0]
                total_salaries = float(row.get("Hugo_Salary", 0)) + float(row.get("Philip_Salary", 0))
                total_cos = sum(float(row.get(COST_COL_MAP[item], 0)) for item in COST_OF_SALES_ITEMS)
                open_balance = float(row.get("Open_Balance", 0))
                paid_bonus = float(row.get("Paid_Bonus", 0))
                target = (total_salaries / 0.4) + total_cos + open_balance

        total_expenses = total_cos + open_balance
        bonus_raw = 0.9 * (((total_rev - total_expenses) * 0.4) - total_salaries)
        net_rev_q = total_rev - total_salaries - total_expenses - paid_bonus

        # YTD costs: sum all quarters within the fiscal year
        ytd_salaries = 0; ytd_expenses = 0; ytd_paid_bonus = 0
        if not costs_df.empty:
            def _in_fy(q_str):
                try:
                    p = pd.Period(q_str, freq='Q')
                    return pd.Timestamp(fy_start) <= p.start_time <= pd.Timestamp(now)
                except: return False
            ytd_c = costs_df[costs_df['Quarter'].apply(_in_fy)]
            if not ytd_c.empty:
                ytd_salaries = pd.to_numeric(ytd_c['Hugo_Salary'], errors='coerce').fillna(0).sum() + \
                               pd.to_numeric(ytd_c['Philip_Salary'], errors='coerce').fillna(0).sum()
                ytd_cos = sum(pd.to_numeric(ytd_c[COST_COL_MAP[i]], errors='coerce').fillna(0).sum()
                              for i in COST_OF_SALES_ITEMS if COST_COL_MAP[i] in ytd_c.columns)
                ytd_ob = pd.to_numeric(ytd_c['Open_Balance'], errors='coerce').fillna(0).sum() if 'Open_Balance' in ytd_c.columns else 0
                ytd_paid_bonus = pd.to_numeric(ytd_c['Paid_Bonus'], errors='coerce').fillna(0).sum() if 'Paid_Bonus' in ytd_c.columns else 0
                ytd_expenses = ytd_cos + ytd_ob
        net_rev_ytd = ytd_rev - ytd_salaries - ytd_expenses - ytd_paid_bonus

        # --- P&L EXPORT BUTTON ---
        _exp_col, _ = st.columns([1, 3])
        with _exp_col:
            try:
                _xl = build_pl_export(df, costs_df)
                st.download_button(
                    "📥 Export P&L to Excel",
                    data=_xl,
                    file_name=f"PL_Monaco_{datetime.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as _e:
                st.error(f"Export error: {_e}")

        st.markdown("### 🏦 Executive Summary")
        kpi1, kpi2, kpi3, kpi4 = st.columns(4)
        kpi1.metric("Quarterly Revenue", f"€{total_rev:,.0f}")
        kpi2.metric("YTD Revenue (Apr–Apr)", f"€{ytd_rev:,.0f}")
        kpi3.metric("Net Revenue (Quarter)", f"€{net_rev_q:,.0f}")
        kpi4.metric("Net Revenue (YTD)", f"€{net_rev_ytd:,.0f}")

        # Bonus card
        st.markdown("")
        _, bonus_col, _ = st.columns([1, 2, 1])
        with bonus_col:
            if bonus_raw >= 0:
                st.markdown(f"""
                <div style="background:linear-gradient(135deg,#1d3557,#457b9d);border-radius:14px;padding:24px 28px;text-align:center;color:white;box-shadow:0 4px 15px rgba(0,0,0,0.15);">
                    <div style="font-size:0.8em;letter-spacing:2px;opacity:0.75;margin-bottom:6px;">BONUS ESTIMATION</div>
                    <div style="font-size:2.6em;font-weight:800;letter-spacing:-1px;">€ {bonus_raw:,.0f}</div>
                    <div style="margin-top:14px;font-size:0.78em;opacity:0.65;">= 0.9 × ((Revenue − Expenses) × 40% − Salaries)</div>
                </div>""", unsafe_allow_html=True)
            else:
                rev_needed = total_expenses + (total_salaries / 0.4)
                gap = rev_needed - total_rev
                pct = min(int((total_rev / rev_needed) * 100), 100) if rev_needed > 0 else 0
                st.markdown(f"""
                <div style="background:linear-gradient(135deg,#6c3a3a,#c0392b);border-radius:14px;padding:24px 28px;text-align:center;color:white;box-shadow:0 4px 15px rgba(0,0,0,0.15);">
                    <div style="font-size:0.8em;letter-spacing:2px;opacity:0.75;margin-bottom:6px;">BONUS ESTIMATION</div>
                    <div style="font-size:2.6em;font-weight:800;">€ 0</div>
                    <div style="background:rgba(255,255,255,0.2);border-radius:6px;height:10px;margin:14px 0 6px;">
                        <div style="background:#f4a261;border-radius:6px;height:10px;width:{pct}%;"></div>
                    </div>
                    <div style="font-size:0.85em;opacity:0.85;">Need <b>€ {gap:,.0f}</b> more revenue to unlock bonus</div>
                    <div style="font-size:0.75em;opacity:0.6;margin-top:4px;">{pct}% of target reached</div>
                </div>""", unsafe_allow_html=True)

        st.divider()
        c1, c2 = st.columns([2, 1])
        with c1:
            rev_by_q = (
                pdf.groupby('Quarter', sort=True)['Rev_EUR']
                .sum()
                .reset_index()
                .rename(columns={'Rev_EUR': 'Revenue (€)'})
            )
            rev_by_q['Cumulative (€)'] = rev_by_q['Revenue (€)'].cumsum()
            total_cumul = rev_by_q['Cumulative (€)'].iloc[-1] if not rev_by_q.empty else 0
            bar_labels = [f"€{v/1000:.0f}K" if v >= 1000 else f"€{v:.0f}" for v in rev_by_q['Revenue (€)']]
            st.markdown(f"<div style='text-align:right;color:#2a9d8f;font-size:1.3em;font-weight:700;margin-bottom:-10px'>€ {total_cumul:,.0f}</div>", unsafe_allow_html=True)
            q_labels  = rev_by_q['Quarter'].tolist()
            q_rev     = rev_by_q['Revenue (€)'].tolist()
            q_cumul   = rev_by_q['Cumulative (€)'].tolist()
            fig_trend = go.Figure()
            fig_trend.add_trace(go.Bar(
                x=q_labels,
                y=q_rev,
                orientation='v',
                name='Quarterly Revenue',
                marker_color='#1d3557',
                text=bar_labels,
                textposition='outside',
                textfont=dict(size=11),
                hovertemplate='<b>%{x}</b><br>Revenue: €%{y:,.0f}<extra></extra>'
            ))
            fig_trend.add_trace(go.Scatter(
                x=q_labels,
                y=q_cumul,
                name='Cumulative',
                mode='lines+markers',
                line=dict(color='#2a9d8f', width=2.5),
                marker=dict(size=7),
                yaxis='y2',
                hovertemplate='<b>%{x}</b><br>Cumulative: €%{y:,.0f}<extra></extra>'
            ))
            fig_trend.update_layout(
                height=370,
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                margin=dict(t=40, b=30, l=60, r=60),
                xaxis=dict(title='', showgrid=False, type='category'),
                yaxis=dict(title='Quarterly (€)', showgrid=True, gridcolor='#eee', tickprefix='€', tickformat=',.0f', side='left'),
                yaxis2=dict(title='Cumulative (€)', tickprefix='€', tickformat=',.0f', overlaying='y', side='right', showgrid=False),
                legend=dict(orientation='h', y=1.12, x=0),
                hovermode='x unified',
                bargap=0.35,
            )
            st.plotly_chart(fig_trend, use_container_width=True)

        with c2:
            if target > 0: st.plotly_chart(create_gauge(total_rev, target), use_container_width=True)
            else: st.info("Add costs for this quarter to see Target Gauge")

        st.divider()
        col_p, col_c = st.columns(2)
        with col_p:
            st.subheader("Revenue by Product")
            p_revs = f_pdf.groupby('Product Type')['Rev_EUR'].sum().sort_values()
            st.plotly_chart(px.bar(p_revs, orientation='h', color_discrete_sequence=['#457b9d']), use_container_width=True)
        with col_c:
            st.subheader("Top Clients")
            c_revs = f_pdf.groupby('Client Name')['Rev_EUR'].sum().sort_values().tail(5)
            st.plotly_chart(px.bar(c_revs, orientation='h', color_discrete_sequence=['#a8dadc']), use_container_width=True)

# --- TAB: COSTS MANAGEMENT ---
with tab_costs:
    st.subheader("Quarterly Financial Entry")
    c_hdr1, c_hdr2 = st.columns(2)
    q_num = c_hdr1.selectbox("Quarter", [1, 2, 3, 4], key="cost_q_sel")
    q_year = c_hdr2.number_input("Year", value=datetime.now().year, key="cost_yr_sel")
    q_key = f"{q_year}Q{q_num}"

    existing = {}
    if not costs_df.empty:
        c_row = costs_df[costs_df['Quarter'] == q_key]
        if not c_row.empty: existing = c_row.iloc[0].to_dict()

    with st.form("detailed_costs_form"):
        h1, h2, h3, h4 = st.columns([2, 1, 1, 1])
        h2.markdown("**Hugo**")
        h3.markdown("**Philip**")
        h4.markdown("**Total**")
        st.divider()

        st.markdown("### Salary & Bonus")
        s_label, s_hugo, s_phil, s_tot = st.columns([2, 1, 1, 1])
        s_label.write("Salaries")
        val_h_sal = s_hugo.number_input("Salaries (H)", label_visibility="collapsed", min_value=0.0, value=float(existing.get("Hugo_Salary", 0.0)))
        val_p_sal = s_phil.number_input("Salaries (P)", label_visibility="collapsed", min_value=0.0, value=float(existing.get("Philip_Salary", 0.0)))
        s_tot.markdown(f"**€ {val_h_sal + val_p_sal:,.2f}**")
        b_label, _, _, b_tot = st.columns([2, 1, 1, 1])
        b_label.write("Paid Bonus")
        val_paid_bonus = b_tot.number_input("Paid Bonus", label_visibility="collapsed", min_value=0.0, value=float(existing.get("Paid_Bonus", 0.0)))
        st.divider()

        st.markdown("### Cost of Sales")
        cos_entries = {}
        for item_name in COST_OF_SALES_ITEMS:
            col_slug = COST_COL_MAP[item_name]
            l_col, h_col, p_col, t_col = st.columns([2, 1, 1, 1])
            l_col.write(item_name)
            h_val = h_col.number_input(f"{item_name}_H", label_visibility="collapsed", min_value=0.0, value=float(existing.get(f"{col_slug}_Hugo", 0.0)))
            p_val = p_col.number_input(f"{item_name}_P", label_visibility="collapsed", min_value=0.0, value=float(existing.get(f"{col_slug}_Philip", 0.0)))
            total_line = h_val + p_val
            t_col.markdown(f"€ {total_line:,.2f}")
            cos_entries[f"{col_slug}_Hugo"] = h_val
            cos_entries[f"{col_slug}_Philip"] = p_val
            cos_entries[col_slug] = total_line
        st.divider()

        st.markdown("### Other Expenses")
        ob_l, ob_h, ob_p, ob_t = st.columns([2, 1, 1, 1])
        ob_l.write("Open Balance")
        val_open_balance = ob_t.number_input("Open Balance", label_visibility="collapsed", min_value=0.0, value=float(existing.get("Open_Balance", 0.0)))
        ob_h.markdown("")
        ob_p.markdown("")
        st.divider()

        total_hugo = val_h_sal + sum(v for k, v in cos_entries.items() if "_Hugo" in k)
        total_philip = val_p_sal + sum(v for k, v in cos_entries.items() if "_Philip" in k)
        grand_total = total_hugo + total_philip + val_open_balance

        f1, f2, f3, f4 = st.columns([2, 1, 1, 1])
        f1.markdown("#### TOTAL PERIOD COSTS")
        f2.markdown(f"**€ {total_hugo:,.2f}**")
        f3.markdown(f"**€ {total_philip:,.2f}**")
        f4.markdown(f"#### € {grand_total:,.2f}")

        if st.form_submit_button("Save Quarterly Financials", use_container_width=True, type="primary"):
            new_row = {"Quarter": q_key, "Hugo_Salary": val_h_sal, "Philip_Salary": val_p_sal, "Total_Salary": val_h_sal + val_p_sal, "Paid_Bonus": val_paid_bonus}
            new_row.update(cos_entries)
            new_row["Open_Balance"] = val_open_balance
            new_row["Total_Costs"] = grand_total + val_paid_bonus

            new_cost_df = pd.DataFrame([new_row])
            updated_costs = new_cost_df if costs_df.empty else pd.concat([costs_df[costs_df['Quarter'] != q_key], new_cost_df], ignore_index=True)

            if save_all_sheets(df, updated_costs, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df):
                st.success(f"Financials for {q_key} saved successfully!")
                time.sleep(1)
                st.rerun()

# --- TAB: ACTION DASHBOARD ---
with tab_dash:
    s_col, q_col, f_col = st.columns([3, 1.5, 1])
    query = s_col.text_input("🔍 Search Blotter...", placeholder="Search Client, ISIN, ID...")
    pending_only = f_col.toggle("Show Pending Only", value=True)

    if not df.empty:
        filtered = df.iloc[::-1].copy()

        # Quarter filter — derive quarters from Trade Date
        _dates = pd.to_datetime(filtered['Trade Date'], format='%d-%b-%y', errors='coerce')
        _quarters = _dates.dt.to_period('Q').dropna().astype(str).unique()
        quarter_options = ["All"] + sorted(_quarters, reverse=True)
        sel_dash_q = q_col.selectbox("📅 Quarter", quarter_options, index=1 if len(_quarters) > 0 else 0, key="dash_q_filter")

        if sel_dash_q != "All":
            filtered = filtered[_dates.dt.to_period('Q').astype(str) == sel_dash_q]

        if query: filtered = filtered[filtered.astype(str).apply(lambda x: x.str.contains(query, case=False)).any(axis=1)]

        # --- BULK FX REFRESH ---
        with st.expander("🔄 Bulk Refresh All FX Rates"):
            st.caption("Fetches the exchange rate for every trade using its trade date, then saves once. Use this to initialise or correct all rates in one go.")
            if st.button("Run Bulk FX Refresh", type="primary", key="bulk_fx_refresh"):
                updated = 0; failed = []
                with st.spinner(f"Fetching rates for {len(df)} trades…"):
                    for _idx, _row in df.iterrows():
                        _ccy = str(_row.get('Currency', '')).strip()
                        _td  = _row.get('Trade Date', '')
                        if not _ccy:
                            continue
                        if _ccy == 'EUR':
                            df.at[_idx, 'Exchange rate'] = 1.0
                            updated += 1
                            continue
                        try:
                            _rate = fetch_rate_for_deal(_ccy, _td)
                            if _rate:
                                df.at[_idx, 'Exchange rate'] = _rate
                                updated += 1
                        except Exception as _e:
                            failed.append(f"{_row.get('Deal ID', _idx)} ({_ccy}): {_e}")
                if save_all_sheets(df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df):
                    msg = f"Updated {updated} / {len(df)} trades."
                    if failed:
                        st.warning(msg + " Failed: " + " | ".join(failed))
                    else:
                        st.success(msg)
                    time.sleep(1); st.rerun()

        for idx, row in filtered.iterrows():
            did = str(row["Deal ID"])
            doc_status = check_docs(did)
            
            _optional = {"Is Internal", "Expersoft ID", "Broker Used", "Broker Name", "Underlying 2", "Underlying 3 & more", "Trade Time", "Redemption Date"}
            missing_info = [c for c in ALL_COLS if c not in _optional and (pd.isna(row.get(c)) or str(row.get(c)).strip() in ["", "nan", "None"])]
            all_docs_present = all(doc_status.values())
            
            if pending_only and not (missing_info or not all_docs_present): 
                continue

            r1, r2, r3, r4 = st.columns([1, 2, 2, 1.2])
            r1.write(f"`{did}`")
            r2.write(f"**{row.get('Client Name', 'N/A')}**")
            
            doc_label = "🟢 Docs" if all_docs_present else "🔴 Docs Missing"
            data_label = "🟢 Data" if not missing_info else f"🔴 {len(missing_info)} Fields Missing"
            r3.write(f"{doc_label} | {data_label}")
            
            if r4.button("Open", key=f"op_{did}"):
                st.session_state.active_deal = None if st.session_state.active_deal == did else did
                st.rerun()

            if st.session_state.active_deal == did:
                with st.container(border=True):
                    # --- THE FINALIZED BUTTON (GREY UNTIL READY) ---
                    is_ready = (not missing_info) and all_docs_present
                    if st.button("🚀 Finalize: Generate MIFIR & Email", key=f"fin_{did}", type="primary", use_container_width=True, disabled=not is_ready):
                        try:
                            updated_payments = finalize_and_send(
                                idx, df, contacts_df, clients_info_df, sub_clients_df,
                                banking_info_df, payments_df
                            )
                            if updated_payments is not None:
                                payments_df = updated_payments
                                save_all_sheets(df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df)
                            st.success("✅ MIFIR Created and Email Drafted!")
                        except Exception as e:
                            st.error(f"Automation Error: {e}")
                    
                    if not is_ready:
                        if missing_info: st.markdown(f"**⚠️ Missing Data:** :red[{', '.join(missing_info)}]")
                        if not all_docs_present:
                            missing_docs_list = [k for k, v in DOC_TYPES.items() if not doc_status[v]]
                            st.markdown(f"**📂 Missing Docs:** :red[{', '.join(missing_docs_list)}]")
                        st.caption("⚠️ The Finalize button will enable once all Data and Documents are complete.")

                    st.divider()

                    d1, d2 = st.columns([3, 2])
                    with d1:
                        curr_row = df[df.index == idx].copy()
                        for c in curr_row.columns:
                            if pd.isna(curr_row.at[idx, c]) or str(curr_row.at[idx, c]) == "nan": curr_row.at[idx, c] = ""

                        config = {
                            "Booker": st.column_config.SelectboxColumn(options=settings.get("Bookers", [])),
                            "Client Name": st.column_config.SelectboxColumn(options=settings.get("Clients", [])),
                            "Issuer": st.column_config.SelectboxColumn(options=settings.get("Issuers", [])),
                            "Sales": st.column_config.SelectboxColumn(options=settings.get("Sales", [])),
                            "Currency": st.column_config.SelectboxColumn(options=settings.get("Currencies", [])),
                            "Exchange rate": st.column_config.NumberColumn(format="%.4f"),
                            "Nominal": st.column_config.NumberColumn(format="%d"),
                            "Retrocession": st.column_config.SelectboxColumn(options=["Yes", "No"]),
                            "Broker Used": st.column_config.SelectboxColumn(options=["Yes", "No"]),
                            "Broker Name": st.column_config.SelectboxColumn(options=settings.get("Brokers", [])),
                            "Product Type": st.column_config.SelectboxColumn(options=settings.get("Product Types", [])),
                        }
                        
                        s1, s2, s3, s4 = st.tabs(["🏛️ Parties", "💰 Econ", "📅 Dates", "🛠️ Execution"])
                        with s1: ed1 = st.data_editor(curr_row[["Booker", "Client Name", "Issuer", "Sales", "ISIN"]], column_config=config, hide_index=True, key=f"e1_{did}")
                        with s2:
                            if st.button("🔄 Refresh Market Rate", key=f"fx_ref_{did}"):
                                try:
                                    new_rate = fetch_rate_for_deal(row.get('Currency'), row.get('Trade Date'))
                                    df.at[idx, 'Exchange rate'] = new_rate
                                    if save_all_sheets(df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df):
                                        st.success(f"Rate updated to {new_rate}")
                                        time.sleep(0.5)
                                        st.rerun()
                                except Exception as e:
                                    st.error(f"FX fetch failed: {e}")
                            ed2 = st.data_editor(curr_row[["Currency", "Nominal", "Exchange rate", "Retrocession", "Fee (%)", "Sales Credit (%)", "Traded Price"]], column_config=config, hide_index=True, key=f"e2_{did}")
                        with s3: ed3 = st.data_editor(curr_row[["Trade Date", "Trade Time", "Issue Date", "Maturity Date", "Redemption Date"]].astype(str), hide_index=True, key=f"e3_{did}")
                        with s4: ed4 = st.data_editor(curr_row[["Product Type", "Underlying 1", "Underlying 2", "Underlying 3 & more", "Broker Used", "Broker Name"]].astype(str), hide_index=True, key=f"e4_{did}")
                        
                        if st.button("💾 Save Changes", key=f"sv_{did}"):
                            for ed in [ed1, ed2, ed3, ed4]:
                                for col in ed.columns:
                                    val = ed.iloc[0][col]
                                    df.at[idx, col] = val if (pd.notna(val) and str(val) != "") else None
                            if save_all_sheets(df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df):
                                st.success("Saved!"); time.sleep(0.5); st.rerun()

                        # Sub-clients section — only for Plurimi clients
                        client_name = row.get('Client Name', '')
                        ci_match = clients_info_df[clients_info_df['Client Name'] == client_name]
                        client_type = str(ci_match.iloc[0].get('Client Type', 'External')).strip() if not ci_match.empty else 'External'
                        if client_type == 'Plurimi':
                            st.divider()
                            st.markdown("#### 👥 Sub-Clients")
                            existing_subs = sub_clients_df[sub_clients_df['Deal_ID'] == did].copy() if not sub_clients_df.empty else pd.DataFrame(columns=SUB_CLIENTS_COLS)
                            if existing_subs.empty:
                                existing_subs = pd.DataFrame([{"Deal_ID": did, "Expersoft_ID": "", "Nominal": 0.0}])
                            edited_subs = st.data_editor(
                                existing_subs[["Expersoft_ID", "Nominal"]],
                                num_rows="dynamic", hide_index=True, use_container_width=True,
                                key=f"subs_{did}",
                                column_config={
                                    "Expersoft_ID": st.column_config.TextColumn(),
                                    "Nominal": st.column_config.NumberColumn(format="%.0f")
                                }
                            )
                            if st.button("💾 Save Sub-Clients", key=f"save_subs_{did}"):
                                edited_subs["Deal_ID"] = did
                                edited_subs = edited_subs[
                                    edited_subs["Expersoft_ID"].notna() &
                                    (edited_subs["Expersoft_ID"].astype(str).str.strip() != "") &
                                    (pd.to_numeric(edited_subs["Nominal"], errors='coerce').fillna(0) > 0)
                                ]
                                other_subs = sub_clients_df[sub_clients_df['Deal_ID'] != did] if not sub_clients_df.empty else pd.DataFrame(columns=SUB_CLIENTS_COLS)
                                new_sub_clients_df = pd.concat([other_subs, edited_subs[SUB_CLIENTS_COLS]], ignore_index=True)
                                if save_all_sheets(df, costs_df, contacts_df, clients_info_df, new_sub_clients_df, banking_info_df, payments_df):
                                    st.success("Sub-clients saved!")
                                    time.sleep(0.5); st.rerun()

                    with d2:
                        st.markdown("#### 📂 Documents")
                        deal_path = find_deal_folder(did)
                        for name, slug in DOC_TYPES.items():
                            exists = doc_status[slug]
                            cn, cb = st.columns([3, 2])
                            cn.write(f"{'✅' if exists else '❌'} {name}")
                            if exists and deal_path:
                                try:
                                    allf = os.listdir(deal_path)
                                    tf = next((f for f in allf if slug in f), None)
                                    if tf: cb.download_button("📥 Get", (deal_path/tf).read_bytes(), file_name=tf, key=f"dl_{did}_{slug}")
                                except: pass
                        st.divider()
                        missing_docs = [n for n, s in DOC_TYPES.items() if not doc_status[s]]
                        if missing_docs:
                            ut = st.selectbox("Upload Type", missing_docs, key=f"sel_{did}")
                            uf = st.file_uploader("Drop file", key=f"up_{did}")
                            if st.button("Confirm Upload", key=f"u_b_{did}"):
                                if uf:
                                    upload_path = get_deal_path(did, row)
                                    os.makedirs(upload_path, exist_ok=True)
                                    with open(upload_path / f"{did}_{DOC_TYPES[ut]}{os.path.splitext(uf.name)[1]}", "wb") as f: f.write(uf.getbuffer())
                                    st.rerun()


# --- TAB: NEW BOOKING ---
with tab_book:
    st.subheader("New Trade Registration")
    next_id_val = f"D{int(str(df['Deal ID'].iloc[-1]).replace('D', '')) + 1:05d}" if not df.empty else "D00001"
    
    with st.form(f"book_form_{st.session_state.form_reset_key}", border=True):
        st.caption(f"Next Available ID: {next_id_val}")
        
        st.markdown("##### 🏛️ Parties & Product")
        c1, c2, c3 = st.columns(3)
        f_booker = c1.selectbox("Booker", [""] + settings.get("Bookers", []))
        f_client = c2.selectbox("Client Name", [""] + settings.get("Clients", []))
        f_issuer = c3.selectbox("Issuer", [""] + settings.get("Issuers", []))
        f_sales = c1.selectbox("Sales", [""] + settings.get("Sales", []))
        f_isin = c2.text_input("ISIN", value="", placeholder="Enter ISIN...")
        f_prod = c3.selectbox("Product Type", [""] + settings.get("Product Types", []))
        
        st.divider()
        
        st.markdown("##### 💰 Economics")
        e1, e2, e3 = st.columns(3)
        f_curr = e1.selectbox("Currency", [""] + settings.get("Currencies", []))
        f_nom = e2.number_input("Nominal", min_value=0.0, value=0.0, step=1000.0)
        f_px = e3.text_input("Traded Price (%)", value="", placeholder="e.g. 100.00%")
        
        e4, e5, e6 = st.columns(3)
        f_ret = e4.selectbox("Retrocession", ["", "No", "Yes"])
        f_fee = e5.number_input("Fee (%)", min_value=0.0, value=0.0, step=0.01)
        f_sc = e6.number_input("Sales Credit (%)", min_value=0.0, value=0.0, step=0.01)
        
        st.divider()
        
        st.markdown("##### 📅 Timeline")
        d1, d2, d3, d4 = st.columns(4)
        f_td = d1.date_input("Trade Date", value=datetime.today())
        f_tt = d2.time_input("Trade Time", value=None)
        f_id = d3.date_input("Issue Date", value=datetime.today())
        f_md = d4.date_input("Maturity Date", value=datetime.today())
        
        st.divider()
        
        st.markdown("##### 🛠️ Underlyings & Brokerage")
        u1, u2, u3 = st.columns(3)
        f_u1 = u1.text_input("Underlying 1", value="")
        f_u2 = u2.text_input("Underlying 2", value="")
        f_u3 = u3.text_input("Underlying 3 & more", value="")
        
        b1, b2 = st.columns(2)
        f_bu = b1.selectbox("Broker Used", ["", "No", "Yes"])
        b_list = [""] + settings.get("Brokers", [])
        f_bn = b2.selectbox("Broker Name", b_list)

        if st.form_submit_button("🚀 Finalize and Book Trade", type="primary", use_container_width=True):
            if not f_booker or not f_client or not f_isin or not f_curr or not f_nom:
                st.error("Please fill in the Booker, Client, ISIN, Currency, and Nominal at a minimum.")
            else:
                fx = get_exchange_rate(f_curr, f_td) or 1.0
                new_row = {
                    "Deal ID": next_id_val, "Booker": f_booker, "Client Name": f_client, "Issuer": f_issuer, 
                    "ISIN": f_isin, "Sales": f_sales, "Currency": f_curr, "Nominal": f_nom, "Exchange rate": fx, 
                    "Retrocession": f_ret, "Fee (%)": f_fee, "Sales Credit (%)": f_sc, "Traded Price": f_px, 
                    "Trade Time": f_tt.strftime("%H:%M:%S") if f_tt else "", 
                    "Trade Date": f_td.strftime("%d-%b-%y"), 
                    "Issue Date": f_id.strftime("%d-%b-%y"), 
                    "Maturity Date": f_md.strftime("%d-%b-%y"), 
                    "Redemption Date": f_md.strftime("%d-%b-%y"), 
                    "Product Type": f_prod, "Underlying 1": f_u1, 
                    "Underlying 2": f_u2, "Underlying 3 & more": f_u3, 
                    "Broker Used": f_bu, "Broker Name": f_bn
                }
                new_df = pd.concat([df, pd.DataFrame([new_row]).astype(object)], ignore_index=True)
                if save_all_sheets(new_df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df):
                    nom_str = f"{int(f_nom):,}" if f_nom else "0"
                    folder_name = f"{next_id_val}_{f_td.strftime('%d.%m.%Y')}_{f_isin}_{nom_str} {f_curr}"
                    os.makedirs(BASE_TRADE_FOLDER / folder_name, exist_ok=True)
                    st.success(f"Deal {next_id_val} successfully recorded!"); reset_form(); time.sleep(1); st.rerun()
                    
# --- TAB: PAYMENTS ---
with tab_pay:
    st.header("💳 Payments")

    pay_left, pay_right = st.columns(2)

    # ── LEFT: Invoices Owed to Plurimi (all deals, issuer → Plurimi) ──
    with pay_left:
        st.subheader("📥 Invoices Owed to Plurimi")
        st.caption("Fees to be received from issuers.")

        OWED_COLS = ["Deal_ID", "ISIN", "Issuer", "Currency", "Nominal",
                     "Issuer Fee (%)", "Issuer Fee Amount", "Issue Date", "Issuer Status"]

        if payments_df.empty:
            st.info("No records yet — finalize a deal to populate.")
        else:
            owed_df = payments_df[
                [c for c in OWED_COLS if c in payments_df.columns]
            ].copy()
            edited_owed = st.data_editor(
                owed_df, hide_index=True, use_container_width=True,
                key="editor_owed",
                column_config={
                    "Issuer Status": st.column_config.SelectboxColumn(options=["Pending", "Received"]),
                }
            )
            if st.button("💾 Save", key="save_owed", type="primary"):
                # Write edited status columns back into payments_df
                payments_df["Issuer Status"] = edited_owed["Issuer Status"].values
                if save_all_sheets(df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df):
                    st.success("Saved!"); time.sleep(1); st.rerun()

    # ── RIGHT: Invoices Due by Plurimi (retro deals only, Plurimi → client) ──
    with pay_right:
        st.subheader("📤 Invoices Due by Plurimi")
        st.caption("Distribution fees to be paid to external clients.")

        DUE_COLS = ["Deal_ID", "ISIN", "Client Name", "Currency", "Nominal",
                    "Retro Fee (%)", "Retro Amount", "Issue Date", "Payment Status"]

        if payments_df.empty:
            st.info("No records yet.")
        else:
            due_df = payments_df[
                payments_df["Payment Status"].astype(str) == "Pending"
            ][[c for c in DUE_COLS if c in payments_df.columns]].copy()

            if due_df.empty:
                st.info("No retrocession deals found.")
            else:
                edited_due = st.data_editor(
                    due_df, hide_index=True, use_container_width=True,
                    key="editor_due",
                    column_config={
                        "Payment Status": st.column_config.SelectboxColumn(options=["Pending", "Paid"]),
                    }
                )
                if st.button("💾 Save", key="save_due", type="primary"):
                    payments_df.loc[
                        payments_df["Payment Status"].astype(str) != "N/A", "Payment Status"
                    ] = edited_due["Payment Status"].values
                    if save_all_sheets(df, costs_df, contacts_df, clients_info_df, sub_clients_df, banking_info_df, payments_df):
                        st.success("Saved!"); time.sleep(1); st.rerun()

                st.divider()
                st.subheader("📄 Generate Invoice PDF")
                deal_options = {
                    f"{r['Deal_ID']} — {r.get('Client Name','')} / {r.get('ISIN','')}": r['Deal_ID']
                    for _, r in due_df.iterrows()
                }
                selected_label = st.selectbox("Select deal", list(deal_options.keys()), key="pay_deal_sel")
                selected_did   = deal_options[selected_label]
                blotter_match  = df[df['Deal ID'].astype(str) == str(selected_did)]

                if not blotter_match.empty:
                    selected_row = blotter_match.iloc[0]
                    cli = str(selected_row.get('Client Name', ''))
                    deal_curr = str(selected_row.get('Currency', ''))
                    company_for_deal = cli
                    if not clients_info_df.empty:
                        ci_row = clients_info_df[clients_info_df['Client Name'] == cli]
                        if not ci_row.empty:
                            company_for_deal = str(ci_row.iloc[0].get('Company', '') or cli)
                    bk_match = pd.DataFrame()
                    if not banking_info_df.empty:
                        bk_match = banking_info_df[
                            (banking_info_df['Company'].astype(str) == company_for_deal) &
                            (banking_info_df['Currency'].astype(str).str.upper() == deal_curr.upper())
                        ]
                        if bk_match.empty:
                            bk_match = banking_info_df[banking_info_df['Company'].astype(str) == company_for_deal]
                    if bk_match.empty:
                        st.warning(f"No banking info for **{company_for_deal}** ({deal_curr}). Add it in Settings → Banking Information.")
                    else:
                        b = bk_match.iloc[0]
                        st.caption(f"{b.get('Salutation','')} {b.get('Contact Name','')} | {b.get('Bank Name','')} | `{b.get('IBAN','')}`")

                    if st.button("📄 Generate Invoice PDF", key="gen_invoice", type="primary"):
                        try:
                            deal_path = get_deal_path(str(selected_row.get('Deal ID','')), selected_row.to_dict())
                            pdf_path  = draft_invoice_pdf(selected_row.to_dict(), banking_info_df, deal_path, clients_info_df)
                            os.startfile(str(pdf_path))
                            st.success(f"Saved: `{pdf_path.name}`")
                        except Exception as e:
                            st.error(f"Invoice generation failed: {e}")


# --- TAB: SETTINGS ---
with tab_settings:
    st.header("Dropdown Management")
    st.info("Values added here are saved directly to the 'Data Settings' sheet.")
    s1, s2 = st.columns(2)
    for i, (cat, items) in enumerate(settings.items()):
        with (s1 if i % 2 == 0 else s2):
            with st.expander(f"⚙️ {cat} ({len(items)})"):
                nv = st.text_input(f"New {cat}", key=f"nv_{cat}")
                if st.button("➕ Add", key=f"ab_{cat}"):
                    if nv and nv not in settings[cat]:
                        settings[cat].append(nv)
                        if save_settings_to_excel(settings):
                            st.success(f"Added {nv}"); time.sleep(0.5); st.rerun()
                st.divider()
                for it in items: st.text(f"• {it}")
    
    st.divider()
    st.header("Issuer Invoicing Contacts")
    st.info("Map default email addresses for MIFIR generation here. This maps directly to the 'Contacts' sheet.")
    
    # Auto-fill missing issuers from Data Settings into Contacts
    current_issuers = settings.get("Issuers", [])
    for iss in current_issuers:
        if iss not in contacts_df['Issuer'].values:
            contacts_df = pd.concat([contacts_df, pd.DataFrame([{"Issuer": iss, "Customer ID": "", "Contact1": "", "Contact2": "", "Contact3": ""}])], ignore_index=True)
    
    # Render Data Editor
    edited_contacts = st.data_editor(contacts_df, hide_index=True, use_container_width=True, key="editor_contacts")
    if st.button("💾 Save Contacts Mapping", type="primary"):
        if save_all_sheets(df, costs_df, edited_contacts, clients_info_df, sub_clients_df, banking_info_df, payments_df):
            st.success("Contacts saved successfully!")
            time.sleep(1)
            st.rerun()

    st.divider()
    st.header("Client Information")
    st.info("Store company details for each client. This maps directly to the 'Clients_Info' sheet.")

    # Auto-fill missing clients from Data Settings into Clients_Info
    current_clients = settings.get("Clients", [])
    for cli in current_clients:
        if cli not in clients_info_df['Client Name'].values:
            clients_info_df = pd.concat([clients_info_df, pd.DataFrame([{
                "Client Name": cli, "Company": "", "Vizibility ID": "", "Expersoft ID": "",
                "Client Type": "External", "Email": "", "Phone": "", "Mobile": ""
            }])], ignore_index=True)
    for col in CLIENTS_INFO_COLS:
        if col not in clients_info_df.columns:
            clients_info_df[col] = "" if col != "Client Type" else "External"

    edited_clients_info = st.data_editor(
        clients_info_df, hide_index=True, use_container_width=True,
        key="editor_clients_info",
        column_config={"Client Type": st.column_config.SelectboxColumn(options=["External", "Individual", "Plurimi"])}
    )
    if st.button("💾 Save Client Information", type="primary", key="save_clients_info"):
        if save_all_sheets(df, costs_df, contacts_df, edited_clients_info, sub_clients_df, banking_info_df, payments_df):
            st.success("Client information saved successfully!")
            time.sleep(1)
            st.rerun()

    st.divider()
    st.header("Banking Information")
    st.info("One row per company per currency. Used to populate the distribution fee invoice PDF.")

    edited_banking = st.data_editor(
        banking_info_df, hide_index=True, use_container_width=True,
        key="editor_banking_info",
        num_rows="dynamic",
    )
    if st.button("💾 Save Banking Information", type="primary", key="save_banking_info"):
        if save_all_sheets(df, costs_df, contacts_df, clients_info_df, sub_clients_df, edited_banking, payments_df):
            st.success("Banking information saved successfully!")
            time.sleep(1)
            st.rerun()